"""Extract data from IMF Q-CRAFT Excel workbook to Parquet files.

Reads the Excel workbook with openpyxl (data_only=True) and extracts:
- Macrofiscal: GDP, revenue, expenditure, debt, etc. per country per year
- Demography: Population by age group and variant per country per year
- Productivity: Productivity levels per country per year
- Climate: GDP loss % per country per scenario per year

Output: 4 Parquet files in data/processed/
"""

from pathlib import Path

import openpyxl
import polars as pl
import pycountry

WORKBOOK_PATH = Path("source-materials/2024_IMF-FAD_Q-CRAFT-Tool-v10.xlsx")
OUTPUT_DIR = Path("data/processed")

# ── Country name → ISO3C mapping ──────────────────────────────────────────────

# Manual overrides for IMF names that don't match pycountry
_MANUAL_ISO3C: dict[str, str] = {
    "Bahamas, The": "BHS",
    "Bolivia": "BOL",
    "Brunei Darussalam": "BRN",
    "Cabo Verde": "CPV",
    "China, People's Republic of": "CHN",
    "Congo, Dem. Rep. of the": "COD",
    "Congo, Republic of": "COG",
    "Côte d'Ivoire": "CIV",
    "Czech Republic": "CZE",
    "Democratic Republic of the Congo": "COD",
    "Dominican Rep.": "DOM",
    "Egypt": "EGY",
    "Egypt, Arab Rep.": "EGY",
    "Eswatini": "SWZ",
    "Gambia, The": "GMB",
    "Hong Kong SAR": "HKG",
    "Iran": "IRN",
    "Iran, Islamic Republic of": "IRN",
    "Korea": "KOR",
    "Korea, Republic of": "KOR",
    "Korea, Rep.": "KOR",
    "Kyrgyz Republic": "KGZ",
    "Lao P.D.R.": "LAO",
    "Macao SAR": "MAC",
    "Micronesia, Fed. States of": "FSM",
    "Moldova": "MDA",
    "North Macedonia": "MKD",
    "OECD members": "OED",
    "Palestine": "PSE",
    "Puerto Rico": "PRI",
    "Republic of Congo": "COG",
    "Russia": "RUS",
    "Russian Federation": "RUS",
    "São Tomé and Príncipe": "STP",
    "Slovak Republic": "SVK",
    "South Korea": "KOR",
    "South Sudan, Republic of": "SSD",
    "St. Kitts and Nevis": "KNA",
    "St. Lucia": "LCA",
    "St. Vincent and the Grenadines": "VCT",
    "Syria": "SYR",
    "Syrian Arab Republic": "SYR",
    "Taiwan Province of China": "TWN",
    "Tanzania": "TZA",
    "Timor-Leste": "TLS",
    "Trinidad and Tobago": "TTO",
    "Turkmenistan": "TKM",
    "Türkiye": "TUR",
    "Venezuela": "VEN",
    "Vietnam": "VNM",
    "Viet Nam": "VNM",
    "West Bank and Gaza": "PSE",
    "Yemen, Republic of": "YEM",
    "Yemen": "YEM",
}

# Names to skip (aggregates, not countries)
_SKIP_NAMES: set[str] = {
    "Africa Eastern and Southern",
    "Africa Western and Central",
    "Arab World",
    "Caribbean small states",
    "Central Europe and the Baltics",
    "Early-demographic dividend",
    "East Asia & Pacific (excluding high income)",
    "Euro area",
    "Europe & Central Asia (excluding high income)",
    "European Union",
    "Fragile and conflict affected situations",
    "Heavily indebted poor countries (HIPC)",
    "High income",
    "Late-demographic dividend",
    "Latin America & Caribbean (excluding high income)",
    "Least developed countries: UN classification",
    "Low & middle income",
    "Low income",
    "Lower middle income",
    "Middle East & North Africa (excluding high income)",
    "Middle income",
    "North America",
    "Not classified",
    "Other small states",
    "Pacific island small states",
    "Small states",
    "South Asia",
    "Sub-Saharan Africa (excluding high income)",
    "Upper middle income",
    "World",
}


def _get_iso3c(country_name: str) -> str | None:
    """Map country name to ISO 3166-1 alpha-3 code."""
    if country_name in _SKIP_NAMES:
        return None
    if country_name in _MANUAL_ISO3C:
        return _MANUAL_ISO3C[country_name]

    # Try pycountry lookup
    try:
        c = pycountry.countries.lookup(country_name)
        return c.alpha_3
    except LookupError:
        pass

    # Try fuzzy search
    try:
        results = pycountry.countries.search_fuzzy(country_name)
        if results:
            return results[0].alpha_3
    except LookupError:
        pass

    return None


def _safe_float(val: object) -> float | None:
    """Convert cell value to float, handling errors and special values."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s in ("n/a", "#VALUE!", "--", "..", "", "…"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


# ── Macrofiscal extraction ────────────────────────────────────────────────────


def _extract_macrofiscal(wb: openpyxl.Workbook) -> pl.DataFrame:
    """Extract macrofiscal data from the raw WEO sections."""
    ws = wb["Macrofiscal"]

    # Section map: label row → (data_start_row, variable_name)
    # Each section has: label row, header row (Country, Subject Descriptor, ...),
    # then ~199 country rows.
    sections = [
        (67, "real_gdp"),  # Real GDP (header at 66)
        (268, "nominal_gdp"),  # Nominal GDP (header at 267)
        (469, "gdp_deflator"),  # GDP Deflator (header at 468)
        (670, "revenue"),  # Revenue (header at 669)
        (871, "expenditure"),  # Expenditure = total (header at 870)
        (1072, "overall_balance"),  # Overall balance (header at 1071)
        (1273, "primary_balance"),  # Primary balance (header at 1272)
        (1474, "debt"),  # Debt (header at 1473)
    ]

    # Get year headers from first section header row
    header_row = list(ws.iter_rows(min_row=66, max_row=66, values_only=True))[0]
    year_start_col = 4  # Column E = 1980
    years = []
    for j in range(year_start_col, len(header_row)):
        val = header_row[j]
        if isinstance(val, (int, float)):
            years.append(int(val))
        else:
            break

    # Filter to 2001-2029 (WEO range used by the engine)
    year_col_indices = {
        y: year_start_col + (y - 1980) for y in years if 2001 <= y <= 2029
    }
    filtered_years = sorted(year_col_indices.keys())

    # Extract each section
    section_data: dict[str, dict[str, dict[int, float | None]]] = {}
    for data_start, var_name in sections:
        section_data[var_name] = {}
        for row in ws.iter_rows(
            min_row=data_start, max_row=data_start + 198, values_only=True
        ):
            country_name = row[0]
            if country_name is None or not isinstance(country_name, str):
                continue
            iso3c = _get_iso3c(country_name)
            if iso3c is None:
                continue
            vals: dict[int, float | None] = {}
            for y in filtered_years:
                col_idx = year_col_indices[y]
                if col_idx < len(row):
                    vals[y] = _safe_float(row[col_idx])
            section_data[var_name][iso3c] = vals

    # Build long-format DataFrame
    rows_out: list[dict] = []
    # Use real_gdp countries as the master list
    for iso3c in sorted(section_data["real_gdp"].keys()):
        for y in filtered_years:
            row_data: dict[str, object] = {"iso3c": iso3c, "years": y}
            for var_name in [s[1] for s in sections]:
                val = section_data.get(var_name, {}).get(iso3c, {}).get(y)
                row_data[var_name] = val
            rows_out.append(row_data)

    df = pl.DataFrame(rows_out)

    # Derive additional columns needed by the engine
    df = df.with_columns(
        # Growth rates from levels
        (
            pl.col("real_gdp") / pl.col("real_gdp").shift(1).over("iso3c") * 100 - 100
        ).alias("real_gdp_growth_percent"),
        (
            pl.col("nominal_gdp") / pl.col("nominal_gdp").shift(1).over("iso3c") * 100
            - 100
        ).alias("nominal_gdp_growth_percent"),
        (
            pl.col("gdp_deflator") / pl.col("gdp_deflator").shift(1).over("iso3c") * 100
            - 100
        ).alias("gdp_deflator_growth_percent"),
        # Primary expenditure = revenue - primary_balance
        (pl.col("revenue") - pl.col("primary_balance")).alias("primary_expenditure"),
        # Interest expenditure = total expenditure - primary expenditure
        (pl.col("expenditure") - (pl.col("revenue") - pl.col("primary_balance"))).alias(
            "interest_expenditure"
        ),
        # Total expenditure (rename)
        pl.col("expenditure").alias("total_expenditure"),
    )

    # Percent of GDP columns
    df = df.with_columns(
        (pl.col("revenue") / pl.col("nominal_gdp") * 100).alias("revenue_percent_gdp"),
        (
            (pl.col("revenue") - pl.col("primary_balance"))
            / pl.col("nominal_gdp")
            * 100
        ).alias("primary_expenditure_percent_gdp"),
        (pl.col("primary_balance") / pl.col("nominal_gdp") * 100).alias(
            "primary_balance_percent_gdp"
        ),
        (pl.col("overall_balance") / pl.col("nominal_gdp") * 100).alias(
            "overall_balance_percent_gdp"
        ),
        (
            (pl.col("expenditure") - (pl.col("revenue") - pl.col("primary_balance")))
            / pl.col("nominal_gdp")
            * 100
        ).alias("interest_expenditure_percent_gdp"),
        # Debt-to-GDP
        (pl.col("debt") / pl.col("nominal_gdp") * 100).alias("debt_to_gdp"),
    )

    # Interest rate = interest_expenditure / debt (same year) * 100
    df = df.with_columns(
        (pl.col("interest_expenditure") / pl.col("debt") * 100).alias(
            "interest_rate_percent"
        ),
    )

    # Add country name lookup (reverse iso3c → name from raw data)
    # Build from the section data
    iso3c_to_name: dict[str, str] = {}
    for row in ws.iter_rows(min_row=67, max_row=265, values_only=True):
        name = row[0]
        if name and isinstance(name, str):
            iso = _get_iso3c(name)
            if iso and iso not in iso3c_to_name:
                iso3c_to_name[iso] = name

    name_df = pl.DataFrame(
        {"iso3c": list(iso3c_to_name.keys()), "country": list(iso3c_to_name.values())}
    )
    df = df.join(name_df, on="iso3c", how="left")

    return df


# ── Demography extraction ─────────────────────────────────────────────────────


def _extract_demography(wb: openpyxl.Workbook) -> pl.DataFrame:
    """Extract demography data from hidden raw data section."""
    ws = wb["Demography"]

    # Hidden data sections (variant, age_group, header_row, data_start_row)
    # Each section: 1 label row (variant + age_group), 1 header row (Country, years),
    # then ~199 country rows
    demo_sections = [
        ("Medium", "15-64", 118, 119, 120),
        ("High", "15-64", 319, 320, 321),
        ("Low", "15-64", 520, 521, 522),
        ("Medium", "Total", 721, 722, 723),
        ("High", "Total", 922, 923, 924),
        ("Low", "Total", 1123, 1124, 1125),
        ("Medium", "65+", 1324, 1325, 1326),
        ("High", "65+", 1525, 1526, 1527),
        ("Low", "65+", 1726, 1727, 1728),
    ]

    # Get year headers from first section
    header_row = list(ws.iter_rows(min_row=119, max_row=119, values_only=True))[0]
    # Header: None, 'Country', 1950, 1951, ...
    year_start_col = 2
    years = []
    for j in range(year_start_col, len(header_row)):
        val = header_row[j]
        if isinstance(val, (int, float)):
            years.append(int(val))
        else:
            break

    rows_out: list[dict] = []
    for variant, age_group, _label_row, _header_row, data_start in demo_sections:
        for row in ws.iter_rows(
            min_row=data_start, max_row=data_start + 198, values_only=True
        ):
            country_name = row[1]  # Country in column B
            if country_name is None or not isinstance(country_name, str):
                continue
            if country_name == "Country":
                continue
            iso3c = _get_iso3c(country_name)
            if iso3c is None:
                continue

            for j, y in enumerate(years):
                col_idx = year_start_col + j
                if col_idx < len(row):
                    val = _safe_float(row[col_idx])
                    if val is not None:
                        rows_out.append(
                            {
                                "iso3c": iso3c,
                                "country": country_name,
                                "years": y,
                                "age_group": age_group,
                                "status": variant,
                                "values": val,
                            }
                        )

    return pl.DataFrame(rows_out)


# ── Productivity extraction ───────────────────────────────────────────────────


def _extract_productivity(wb: openpyxl.Workbook) -> pl.DataFrame:
    """Extract productivity level data from hidden raw data section."""
    ws = wb["Productivity"]

    # Hidden data: row 62 = header, rows 63+ = country data
    header_row = list(ws.iter_rows(min_row=62, max_row=62, values_only=True))[0]
    # Header: 'Country Name', 1991, 1992, ...
    year_start_col = 1
    years = []
    for j in range(year_start_col, len(header_row)):
        val = header_row[j]
        if isinstance(val, (int, float)):
            years.append(int(val))
        else:
            break

    rows_out: list[dict] = []
    for row in ws.iter_rows(min_row=63, max_row=300, values_only=True):
        country_name = row[0]
        if country_name is None or not isinstance(country_name, str):
            continue
        iso3c = _get_iso3c(country_name)
        if iso3c is None:
            continue

        for j, y in enumerate(years):
            col_idx = year_start_col + j
            if col_idx < len(row):
                val = _safe_float(row[col_idx])
                if val is not None:
                    rows_out.append(
                        {
                            "iso3c": iso3c,
                            "years": y,
                            "productivity_level": val,
                        }
                    )

    return pl.DataFrame(rows_out)


# ── Climate extraction ────────────────────────────────────────────────────────


def _extract_climate(wb: openpyxl.Workbook) -> pl.DataFrame:
    """Extract climate GDP loss data from Climate Database sheet."""
    ws = wb["Climate Database"]

    # Scenario sections: (scenario_name, header_row, data_start_row)
    climate_sections = [
        ("Paris", 25, 26),
        ("Moderate", 225, 226),
        ("High", 425, 426),
        ("Hot", 625, 626),
        ("Hot_Adapted", 825, 826),
        ("Hot_Unadapted", 1025, 1026),
    ]

    # Get year headers from first section header row
    header_row = list(ws.iter_rows(min_row=25, max_row=25, values_only=True))[0]
    year_start_col = 2  # Column C
    years = []
    for j in range(year_start_col, len(header_row)):
        val = header_row[j]
        if isinstance(val, (int, float)):
            years.append(int(val))
        else:
            break

    rows_out: list[dict] = []
    for scenario, _header, data_start in climate_sections:
        for row in ws.iter_rows(
            min_row=data_start, max_row=data_start + 198, values_only=True
        ):
            country_name = row[1]  # Country in column B
            if country_name is None or not isinstance(country_name, str):
                continue
            iso3c = _get_iso3c(country_name)
            if iso3c is None:
                continue

            for j, y in enumerate(years):
                col_idx = year_start_col + j
                if col_idx < len(row):
                    val = _safe_float(row[col_idx])
                    rows_out.append(
                        {
                            "iso3c": iso3c,
                            "climate_scenario": scenario,
                            "years": y,
                            "gdp_loss_percent": val if val is not None else 0.0,
                        }
                    )

    return pl.DataFrame(rows_out)


# ── Main ──────────────────────────────────────────────────────────────────────


def main() -> None:
    print(f"Loading workbook: {WORKBOOK_PATH}")
    wb = openpyxl.load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Phase 1: Macrofiscal
    print("\n── Extracting Macrofiscal ──")
    df_macro = _extract_macrofiscal(wb)
    uga = df_macro.filter(pl.col("iso3c") == "UGA")
    uga_2009 = uga.filter(pl.col("years") == 2009)
    print(f"  Countries: {df_macro['iso3c'].n_unique()}")
    print(f"  Years: {df_macro['years'].min()}-{df_macro['years'].max()}")
    print(f"  Uganda 2009 nominal_gdp: {uga_2009['nominal_gdp'][0]}")
    print(f"  Uganda 2009 real_gdp: {uga_2009['real_gdp'][0]}")
    df_macro.write_parquet(OUTPUT_DIR / "macrofiscal.parquet")
    print(f"  Written: {OUTPUT_DIR / 'macrofiscal.parquet'}")

    # Phase 2: Demography
    print("\n── Extracting Demography ──")
    df_demo = _extract_demography(wb)
    uga_demo = df_demo.filter(
        (pl.col("iso3c") == "UGA")
        & (pl.col("status") == "Medium")
        & (pl.col("age_group") == "15-64")
        & (pl.col("years") == 2009)
    )
    print(f"  Countries: {df_demo['iso3c'].n_unique()}")
    print(f"  Years: {df_demo['years'].min()}-{df_demo['years'].max()}")
    print(f"  Uganda Medium 15-64 2009: {uga_demo['values'][0]}")
    df_demo.write_parquet(OUTPUT_DIR / "demography.parquet")
    print(f"  Written: {OUTPUT_DIR / 'demography.parquet'}")

    # Phase 3: Productivity
    print("\n── Extracting Productivity ──")
    df_prod = _extract_productivity(wb)
    uga_prod = df_prod.filter((pl.col("iso3c") == "UGA") & (pl.col("years") == 2009))
    print(f"  Countries: {df_prod['iso3c'].n_unique()}")
    print(f"  Years: {df_prod['years'].min()}-{df_prod['years'].max()}")
    if len(uga_prod) > 0:
        print(f"  Uganda 2009 level: {uga_prod['productivity_level'][0]}")
    df_prod.write_parquet(OUTPUT_DIR / "productivity.parquet")
    print(f"  Written: {OUTPUT_DIR / 'productivity.parquet'}")

    # Phase 4: Climate
    print("\n── Extracting Climate ──")
    df_climate = _extract_climate(wb)
    uga_climate = df_climate.filter(
        (pl.col("iso3c") == "UGA")
        & (pl.col("climate_scenario") == "Paris")
        & (pl.col("years") == 2015)
    )
    print(f"  Countries: {df_climate['iso3c'].n_unique()}")
    print(f"  Years: {df_climate['years'].min()}-{df_climate['years'].max()}")
    print(f"  Scenarios: {df_climate['climate_scenario'].unique().sort().to_list()}")
    if len(uga_climate) > 0:
        print(f"  Uganda Paris 2015 gdp_loss: {uga_climate['gdp_loss_percent'][0]}")
    df_climate.write_parquet(OUTPUT_DIR / "climate.parquet")
    print(f"  Written: {OUTPUT_DIR / 'climate.parquet'}")

    wb.close()

    print("\n── Summary ──")
    for name, df in [
        ("macrofiscal", df_macro),
        ("demography", df_demo),
        ("productivity", df_prod),
        ("climate", df_climate),
    ]:
        n = df["iso3c"].n_unique()
        print(f"  {name:14s} {len(df)} rows, {n} countries")
    print("\nDone!")


if __name__ == "__main__":
    main()
