"""Phase 0: Workbook Discovery — map cell references and country names."""

import json
import logging
import os
import sys
from pathlib import Path

import openpyxl

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

PROJECT_ROOT = Path(__file__).resolve().parents[2]
WORKBOOK = PROJECT_ROOT / "source-materials" / "2024_IMF-FAD_Q-CRAFT-Tool-v10.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "verification-logs"
OUTPUT_DIR.mkdir(exist_ok=True)


def build_country_name_map(wb_formula):
    """Build ISO3 ↔ workbook name mapping from the validation source range."""
    # Read country names from Macrofiscal!A67:A264
    ws = wb_formula["Macrofiscal"]
    excel_names = []
    for row in range(67, 265):
        val = ws.cell(row=row, column=1).value
        if val and isinstance(val, str) and val.strip():
            excel_names.append(val.strip())

    logger.info(f"Found {len(excel_names)} country names in Excel validation range")

    # Build ISO3 mapping using the extraction script's approach
    try:
        import pycountry
    except ImportError:
        logger.warning("pycountry not available, using engine country list for mapping")
        pycountry = None

    # Manual overrides from extract_excel_data.py
    _MANUAL_ISO3C = {
        "Afghanistan, Islamic Republic of": "AFG",
        "Bahamas, The": "BHS",
        "Bolivia": "BOL",
        "Brunei Darussalam": "BRN",
        "Cabo Verde": "CPV",
        "China, People's Republic of": "CHN",
        "Congo, Dem. Rep. of the": "COD",
        "Congo, Republic of": "COG",
        "Democratic Republic of the Congo": "COD",
        "Republic of Congo": "COG",
        "Côte d'Ivoire": "CIV",
        "Czech Republic": "CZE",
        "Egypt, Arab Rep.": "EGY",
        "Egypt": "EGY",
        "Equatorial Guinea, Republic of": "GNQ",
        "Eswatini": "SWZ",
        "Gambia, The": "GMB",
        "Guinea-Bissau": "GNB",
        "Hong Kong SAR": "HKG",
        "Iran, Islamic Republic of": "IRN",
        "Iran": "IRN",
        "Korea, Rep.": "KOR",
        "Korea": "KOR",
        "Kosovo": "XKX",
        "Kyrgyz Republic": "KGZ",
        "Lao P.D.R.": "LAO",
        "Lao People's Democratic Republic": "LAO",
        "Macao SAR": "MAC",
        "Marshall Islands, Republic of": "MHL",
        "Micronesia, Fed. States of": "FSM",
        "Moldova": "MDA",
        "Mozambique, Republic of": "MOZ",
        "Myanmar": "MMR",
        "North Macedonia": "MKD",
        "Palau, Republic of": "PLW",
        "Russia": "RUS",
        "Russian Federation": "RUS",
        "São Tomé and Príncipe": "STP",
        "Serbia, Republic of": "SRB",
        "Serbia": "SRB",
        "Slovak Republic": "SVK",
        "St. Kitts and Nevis": "KNA",
        "St. Lucia": "LCA",
        "St. Vincent and the Grenadines": "VCT",
        "Somalia, Federal Republic of": "SOM",
        "South Sudan, Republic of": "SSD",
        "South Sudan": "SSD",
        "Syria": "SYR",
        "Syrian Arab Republic": "SYR",
        "Taiwan Province of China": "TWN",
        "Tanzania": "TZA",
        "Timor-Leste": "TLS",
        "Türkiye": "TUR",
        "Turkey": "TUR",
        "Venezuela": "VEN",
        "Venezuela, Republica Bolivariana de": "VEN",
        "Vietnam": "VNM",
        "Viet Nam": "VNM",
        "West Bank and Gaza": "PSE",
        "Yemen, Republic of": "YEM",
        "Yemen": "YEM",
    }

    _SKIP_NAMES = {
        "World", "OECD members", "Euro area", "High income",
        "Low income", "Lower middle income", "Upper middle income",
        "Middle income", "Advanced economies", "Emerging market and developing economies",
        "Sub-Saharan Africa", "Latin America and the Caribbean",
        "Middle East and Central Asia", "Emerging and Developing Asia",
        "Emerging and Developing Europe", "East Asia and Pacific",
        "Europe and Central Asia", "South Asia",
        "ASEAN-5", "Commonwealth of Independent States",
        "European Union", "G7", "G20", "Major advanced economies (G7)",
        "Other advanced economies", "Small states", "Pacific island small states",
        "Caribbean small states", "Other small states",
        "Heavily indebted poor countries", "Low-income developing countries",
    }

    name_map = {}  # ISO3 -> excel_name
    unmapped = []

    for name in excel_names:
        if name in _SKIP_NAMES:
            continue

        # Try manual override first
        if name in _MANUAL_ISO3C:
            name_map[_MANUAL_ISO3C[name]] = name
            continue

        # Try pycountry
        if pycountry:
            try:
                c = pycountry.countries.lookup(name)
                name_map[c.alpha_3] = name
                continue
            except LookupError:
                pass
            try:
                matches = pycountry.countries.search_fuzzy(name)
                if matches:
                    name_map[matches[0].alpha_3] = name
                    continue
            except LookupError:
                pass

        unmapped.append(name)

    if unmapped:
        logger.warning(f"Unmapped country names: {unmapped}")

    return name_map, excel_names, unmapped


def discover_output_columns(wb_formula):
    """Find output table structure on Output Baseline sheet."""
    ws = wb_formula["Output Baseline"]

    # The output table has years as rows and metrics as columns
    # Scan for header row — look for "Year" or year numbers
    output_info = {}

    # Check first few rows for headers
    for row in range(1, 20):
        for col in range(1, 30):
            val = ws.cell(row=row, column=col).value
            if val and isinstance(val, str):
                val_lower = val.lower().strip()
                if "year" in val_lower:
                    output_info["header_row"] = row
                    output_info["year_col"] = col
                elif "debt" in val_lower and "gdp" in val_lower:
                    output_info["debt_to_gdp_col"] = col
                    output_info["debt_to_gdp_header"] = val
                elif "revenue" in val_lower and "gdp" in val_lower:
                    output_info["revenue_pct_gdp_col"] = col
                elif "primary" in val_lower and "expenditure" in val_lower:
                    output_info["primary_exp_pct_gdp_col"] = col
                elif "primary" in val_lower and "balance" in val_lower:
                    output_info["primary_bal_pct_gdp_col"] = col
                elif "interest" in val_lower and "expenditure" in val_lower:
                    output_info["interest_exp_pct_gdp_col"] = col
                elif "overall" in val_lower and "balance" in val_lower:
                    output_info["overall_bal_pct_gdp_col"] = col

    logger.info(f"Output Baseline structure: {output_info}")
    return output_info


def discover_input_cells(wb_data):
    """Read current (default) values from Dashboard input cells."""
    ws = wb_data["Dashboard"]

    # Correct cell mapping (C17 is demography, C34 is debt_target)
    excel_defaults = {
        "demography_variant": ws["C17"].value,
        "productivity_start": ws["C20"].value,
        "productivity_end": ws["C21"].value,
        "inflation_start": ws["C24"].value,
        "inflation_end": ws["C25"].value,
        "interest_rate_mode": ws["C28"].value,
        "real_interest_rate": ws["C29"].value,
        "fiscal_rule": ws["C33"].value,
        "debt_target": ws["C34"].value,
        "expenditure_rigidity": ws["C38"].value,
    }

    logger.info(f"Excel defaults: {excel_defaults}")
    return excel_defaults, {}


def discover_sheets(wb):
    """List all sheet names."""
    sheets = wb.sheetnames
    logger.info(f"Sheets ({len(sheets)}): {sheets}")
    return sheets


def check_workbook_properties(wb):
    """Check for VBA macros and external links."""
    has_vba = wb.vba_archive is not None if hasattr(wb, 'vba_archive') else False
    has_external_links = False
    try:
        if hasattr(wb, '_external_links') and wb._external_links:
            has_external_links = True
    except Exception:
        pass

    logger.info(f"VBA macros: {has_vba}, External links: {has_external_links}")
    return has_vba, has_external_links


def main():
    logger.info(f"Opening workbook: {WORKBOOK}")

    # Open with formulas (to map structure)
    wb_formula = openpyxl.load_workbook(str(WORKBOOK), data_only=False)
    sheets = discover_sheets(wb_formula)
    has_vba, has_external_links = check_workbook_properties(wb_formula)

    # Build country name mapping
    name_map, all_excel_names, unmapped = build_country_name_map(wb_formula)
    logger.info(f"Mapped {len(name_map)} countries, {len(unmapped)} unmapped")

    # Discover output structure
    output_info = discover_output_columns(wb_formula)

    wb_formula.close()

    # Open with data_only to read cached values
    wb_data = openpyxl.load_workbook(str(WORKBOOK), data_only=True)
    excel_defaults, additional_cells = discover_input_cells(wb_data)

    # Check cached country
    cached_country = wb_data["Dashboard"]["C12"].value
    logger.info(f"Cached country: {cached_country}")

    wb_data.close()

    # Engine defaults comparison
    from qcraft_engine.constants import DEFAULTS as ENGINE_DEFAULTS

    engine_defaults_comparison = {
        "debt_target": {
            "excel": excel_defaults.get("debt_target"),
            "python": ENGINE_DEFAULTS["debt_target"],
            "match": excel_defaults.get("debt_target") == ENGINE_DEFAULTS["debt_target"],
        },
        "fiscal_rule": {
            "excel": excel_defaults.get("fiscal_rule"),
            "python": ENGINE_DEFAULTS["fiscal_rule"],
            "match": excel_defaults.get("fiscal_rule") == ENGINE_DEFAULTS["fiscal_rule"],
        },
        "expenditure_rigidity": {
            "excel": excel_defaults.get("expenditure_rigidity"),
            "python": ENGINE_DEFAULTS["expenditure_rigidity"],
            "match": excel_defaults.get("expenditure_rigidity") == ENGINE_DEFAULTS["expenditure_rigidity"],
        },
        "interest_rate_mode": {
            "excel": excel_defaults.get("interest_rate_mode"),
            "python": ENGINE_DEFAULTS["interest_rate_mode"],
            "match": excel_defaults.get("interest_rate_mode") == ENGINE_DEFAULTS["interest_rate_mode"],
        },
        "inflation_start": {
            "excel": 3.5,  # Known from workbook
            "python": ENGINE_DEFAULTS["inflation_start"],
            "match": False,  # Excel=3.5, Python=5.0
        },
    }

    # Get engine country list for cross-reference
    from qcraft_engine.data_loader import load_parquet_data, get_country_list

    data = load_parquet_data()
    engine_countries = get_country_list(data)
    engine_iso3_set = {c["iso3c"] for c in engine_countries}
    engine_name_map = {c["iso3c"]: c["country"] for c in engine_countries}

    # Cross-reference
    excel_only = set(name_map.keys()) - engine_iso3_set
    engine_only = engine_iso3_set - set(name_map.keys())
    both = set(name_map.keys()) & engine_iso3_set

    logger.info(f"Countries in both: {len(both)}, Excel only: {len(excel_only)}, Engine only: {len(engine_only)}")

    config = {
        "country_selector_cell": "Dashboard!C12",
        "country_name_map": name_map,
        "engine_country_map": engine_name_map,
        "all_excel_names": all_excel_names,
        "unmapped_names": unmapped,
        "countries_both_systems": sorted(both),
        "countries_excel_only": sorted(excel_only),
        "countries_engine_only": sorted(engine_only),
        "input_cells": {
            "country_selector": "Dashboard!C12",
            "demography_variant": "Dashboard!C17",
            "productivity_start": "Dashboard!C20",
            "productivity_end": "Dashboard!C21",
            "inflation_start": "Dashboard!C24",
            "inflation_end": "Dashboard!C25",
            "interest_rate_mode": "Dashboard!C28",
            "fiscal_rule": "Dashboard!C33",
            "debt_target": "Dashboard!C34",
            "expenditure_rigidity": "Dashboard!C38",
        },
        "excel_defaults": excel_defaults,
        "additional_dashboard_cells": additional_cells,
        "engine_defaults_comparison": engine_defaults_comparison,
        "output_info": output_info,
        "output_sheets": {
            "baseline_calc": "Baseline",
            "baseline_summary": "Output Baseline",
            "scenario_calc": {
                "Paris": "Paris",
                "Moderate": "Moderate",
                "Hot": "Hot",
                "Hot_Adapted": "Hot Adapted",
                "Hot_Unadapted": "Hot Unadapted",
                "High": "High",
            },
            "scenario_summary": "Output Scenarios",
        },
        "cached_country": cached_country,
        "weo_max_year": 2029,
        "has_vba_macros": has_vba,
        "has_external_links": has_external_links,
        "check_years": [2030, 2040, 2050, 2070, 2099],
        "sheets": sheets,
    }

    # Serialize — handle non-JSON types
    def make_serializable(obj):
        if isinstance(obj, set):
            return sorted(obj)
        if isinstance(obj, (int, float, str, bool, type(None))):
            return obj
        if isinstance(obj, dict):
            return {k: make_serializable(v) for k, v in obj.items()}
        if isinstance(obj, (list, tuple)):
            return [make_serializable(i) for i in obj]
        return str(obj)

    config = make_serializable(config)

    out_path = OUTPUT_DIR / "phase0_config.json"
    with open(out_path, "w") as f:
        json.dump(config, f, indent=2, ensure_ascii=False)

    logger.info(f"Phase 0 config saved to {out_path}")
    logger.info(f"Total testable countries (in both systems): {len(both)}")

    return config


if __name__ == "__main__":
    main()
